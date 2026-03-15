"""
output.py — Filtering, JSON output, and XLSX export for scraped data.
"""

import json
import os
import re
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def filter_with_phone(listings: list[dict]) -> list[dict]:
    """
    Filter out listings that do not have a valid phone number.

    Args:
        listings: Raw list of extracted business data dicts.

    Returns:
        Filtered list containing only entries with a non-empty phone field.
    """
    filtered = [
        entry for entry in listings
        if entry.get("phone") and entry["phone"].strip()
    ]
    removed = len(listings) - len(filtered)
    logger.info(
        "Filtered: %d with phone, %d removed (no phone)",
        len(filtered), removed,
    )
    return filtered


def _slugify(text: str, max_len: int = 40) -> str:
    """Convert a query string into a filesystem-safe slug."""
    slug = text.lower().strip()
    slug = re.sub(r'[àáâã]', 'a', slug)
    slug = re.sub(r'[éêë]', 'e', slug)
    slug = re.sub(r'[íîï]', 'i', slug)
    slug = re.sub(r'[óôõ]', 'o', slug)
    slug = re.sub(r'[úûü]', 'u', slug)
    slug = re.sub(r'[ç]', 'c', slug)
    slug = re.sub(r'[^a-z0-9]+', '_', slug)
    slug = slug.strip('_')
    return slug[:max_len]


def make_filename(query: str, ext: str = "json") -> str:
    """
    Generate a descriptive filename from a query.

    Format: {query_slug}_{YYYYMMDD_HHMMSS}.{ext}
    Example: clinicas_em_sao_paulo_20260315_170500.json
    """
    slug = _slugify(query)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{slug}_{timestamp}.{ext}"


def save_to_json(
    data: list[dict],
    filename: str = None,
    output_dir: str = "./output",
    query: str = None,
) -> str:
    """
    Save the scraped data to a JSON file.

    If no filename is given, generates one from the query.

    Args:
        data: List of business data dicts to save.
        filename: Output filename (auto-generated if None).
        output_dir: Directory for the output file.
        query: Original search query (used for filename + metadata).

    Returns:
        Absolute path to the saved file.
    """
    os.makedirs(output_dir, exist_ok=True)

    if filename is None:
        filename = make_filename(query or "results")

    filepath = os.path.join(output_dir, filename)

    # Build the output structure with metadata
    output = {
        "metadata": {
            "query": query,
            "scraped_at": datetime.now().isoformat(),
            "total_results": len(data),
        },
        "results": data,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    abs_path = os.path.abspath(filepath)
    logger.info("Saved %d results to %s", len(data), abs_path)
    return abs_path


def export_to_xlsx(
    data: list[dict],
    filepath: str,
    query: str = None,
) -> str:
    """
    Export scraped data to an XLSX spreadsheet.

    Args:
        data: List of business data dicts.
        filepath: Full path for the .xlsx file.
        query: Original search query (for sheet title).

    Returns:
        Absolute path to the saved file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = _slugify(query or "results", max_len=31)

    # --- Header style ---
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # --- Headers ---
    headers = ["#", "Name", "Phone", "Address", "Website", "Rating", "Reviews"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows ---
    data_font = Font(name="Inter", size=10)
    for i, entry in enumerate(data, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = data_font
        ws.cell(row=row, column=2, value=entry.get("name", "")).font = data_font
        ws.cell(row=row, column=3, value=entry.get("phone", "")).font = data_font
        ws.cell(row=row, column=4, value=entry.get("address", "")).font = data_font
        ws.cell(row=row, column=5, value=entry.get("website", "")).font = data_font
        ws.cell(row=row, column=6, value=entry.get("rating", "")).font = data_font
        ws.cell(row=row, column=7, value=entry.get("reviews", "")).font = data_font

        # Light border on each row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    ws.auto_filter.ref = f"A1:G{len(data) + 1}"

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    wb.save(filepath)

    abs_path = os.path.abspath(filepath)
    logger.info("Exported %d results to XLSX: %s", len(data), abs_path)
    return abs_path
